#!/usr/bin/env python3
"""Portable FSRS learning database for the adaptive-study-coach skill."""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import os
import re
import sqlite3
import sys
import uuid
import zipfile
from collections import defaultdict, deque
from collections.abc import Iterable, Sequence
from datetime import datetime, timedelta, timezone
from importlib import metadata
from pathlib import Path
from typing import Any, NoReturn

from path_guard import PathBoundaryError, resolve_inside

DEPENDENCY_IMPORT_ERROR: Exception | None = None
try:
    import pypdf.filters
    from fsrs import Card, Rating, Scheduler
    from openpyxl import load_workbook
    from pypdf import PdfReader

    # Text extraction never needs pypdf's optional external JBIG2 decoder.
    pypdf.filters.JBIG2DEC_BINARY = None
except Exception as exc:  # pragma: no cover - exercised by subprocess failure tests.
    DEPENDENCY_IMPORT_ERROR = exc


SCHEMA_VERSION = "2"
SQLITE_BUSY_TIMEOUT_MS = 15_000
DEFAULT_HOME_NAME = "adaptive-study-data"
LOCAL_TZ = timezone(timedelta(hours=8), name="Asia/Shanghai")
ITEM_TYPES = {"term", "concept", "qa", "cloze", "problem", "procedure"}
MAX_IMPORT_BYTES = 50 * 1024 * 1024
MAX_PDF_PAGES = 500
MAX_EXPANDED_DOCUMENT_BYTES = 200 * 1024 * 1024
MAX_EXTRACTED_TEXT_CHARS = 5_000_000

DEFAULT_CONFIG: dict[str, Any] = {
    "desired_retention": 0.90,
    "daily_minutes": 20,
    "daily_question_limit": 30,
    "weekly_questions": 12,
    "weekly_minutes": 15,
    "monthly_questions": 30,
    "monthly_minutes": 30,
    "auto_import_confidence": 0.75,
    "max_remediation_attempts_per_item": 3,
    "max_remediation_attempts_per_session": 12,
    "backup_retention_count": 30,
    "backup_max_total_mb": 2048,
    "timezone": "Asia/Shanghai",
    "enable_fuzzing": True,
}

HEADER_ALIASES = {
    "subject": "subject",
    "学科": "subject",
    "topic": "topic",
    "主题": "topic",
    "知识点": "topic",
    "章节": "topic",
    "type": "type",
    "类型": "type",
    "题型": "type",
    "prompt": "prompt",
    "问题": "prompt",
    "题目": "prompt",
    "正面": "prompt",
    "单词": "prompt",
    "term": "prompt",
    "answer": "answer",
    "答案": "answer",
    "释义": "answer",
    "背面": "answer",
    "aliases": "aliases",
    "同义答案": "aliases",
    "可接受答案": "aliases",
    "tags": "tags",
    "标签": "tags",
    "source": "source",
    "来源": "source",
    "confidence": "confidence",
    "置信度": "confidence",
    "ocr_confidence": "ocr_confidence",
    "ocr置信度": "ocr_confidence",
    "content_confidence": "content_confidence",
    "内容置信度": "content_confidence",
    "content_verified": "content_verified",
    "内容已确认": "content_verified",
}

TYPE_ALIASES = {
    "术语": "term",
    "单词": "term",
    "概念": "concept",
    "问答": "qa",
    "填空": "cloze",
    "计算": "problem",
    "问题": "problem",
    "步骤": "procedure",
    "流程": "procedure",
}


class StudyError(RuntimeError):
    """A user-actionable manager error."""


class JsonArgumentParser(argparse.ArgumentParser):
    """Convert invalid command lines into the manager's JSON error envelope."""

    def error(self, message: str) -> NoReturn:
        raise StudyError(f"命令参数错误：{message}")


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def iso_utc(value: datetime | None = None) -> str:
    current = value or utc_now()
    if current.tzinfo is None:
        current = current.replace(tzinfo=timezone.utc)
    return current.astimezone(timezone.utc).isoformat()


def parse_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    parsed = datetime.fromisoformat(value)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def local_display(value: str | datetime | None) -> str | None:
    if value is None:
        return None
    parsed = parse_dt(value) if isinstance(value, str) else value
    if parsed is None:
        return None
    return parsed.astimezone(LOCAL_TZ).strftime("%Y-%m-%d %H:%M")


def end_of_local_day_utc(now: datetime | None = None) -> datetime:
    current = (now or utc_now()).astimezone(LOCAL_TZ)
    end = current.replace(hour=23, minute=59, second=59, microsecond=999999)
    return end.astimezone(timezone.utc)


def workspace_root_from_script() -> Path:
    resolved = Path(__file__).resolve()
    for parent in resolved.parents:
        if parent.name == ".agents":
            return parent.parent.resolve()
    raise StudyError("技能不在工作区的 .agents\\skills 目录中。")


def resolve_home(raw_home: str | None) -> tuple[Path, Path]:
    workspace = workspace_root_from_script()
    candidate = raw_home if raw_home else workspace / DEFAULT_HOME_NAME
    try:
        home = resolve_inside(
            workspace,
            candidate,
            must_exist=False,
            allow_root=False,
            label="学习目录",
        )
    except PathBoundaryError as exc:
        raise StudyError(str(exc)) from exc
    return workspace, home


def resolve_workspace_input(path: str | Path, label: str = "导入文件") -> Path:
    workspace = workspace_root_from_script()
    try:
        resolved = resolve_inside(
            workspace,
            path,
            must_exist=True,
            allow_root=False,
            label=label,
        )
    except PathBoundaryError as exc:
        raise StudyError(str(exc)) from exc
    if not resolved.is_file():
        raise StudyError(f"{label}不是普通文件：{resolved}")
    return resolved


def ensure_dirs(home: Path) -> None:
    workspace = workspace_root_from_script()
    try:
        safe_home = resolve_inside(
            workspace,
            home,
            must_exist=False,
            allow_root=False,
            label="学习目录",
        )
    except PathBoundaryError as exc:
        raise StudyError(str(exc)) from exc
    safe_home.mkdir(parents=True, exist_ok=True)
    try:
        safe_home = resolve_inside(
            workspace,
            safe_home,
            must_exist=True,
            allow_root=False,
            label="学习目录",
        )
    except PathBoundaryError as exc:
        raise StudyError(str(exc)) from exc
    for relative in (
        "data",
        "reports",
        "backups",
        "imports",
        "cache/temp",
        "cache/pip",
        "cache/pycache",
        "audit",
    ):
        directory = safe_home / relative
        directory.mkdir(parents=True, exist_ok=True)
        try:
            resolve_inside(
                safe_home,
                directory,
                must_exist=True,
                allow_root=False,
                label=f"输出目录 {relative}",
            )
        except PathBoundaryError as exc:
            raise StudyError(str(exc)) from exc


def atomic_write_text(path: Path, content: str, encoding: str = "utf-8") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    temp_path.write_text(content, encoding=encoding)
    os.replace(temp_path, path)


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_config(home: Path, create: bool = False) -> dict[str, Any]:
    path = home / "config.json"
    if not path.exists():
        if not create:
            raise StudyError("学习系统尚未初始化，请先运行 init。")
        atomic_write_text(path, json.dumps(DEFAULT_CONFIG, ensure_ascii=False, indent=2) + "\n")
        return dict(DEFAULT_CONFIG)
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise StudyError(f"配置文件无法读取：{exc}") from exc
    merged = dict(DEFAULT_CONFIG)
    merged.update(loaded)
    return merged


def db_path(home: Path) -> Path:
    return home / "data" / "study.db"


def begin_immediate(connection: sqlite3.Connection) -> None:
    """Acquire SQLite's reserved writer lock before reading state that will be changed."""
    try:
        connection.execute("BEGIN IMMEDIATE")
    except sqlite3.OperationalError as exc:
        message = str(exc).casefold()
        if "locked" in message or "busy" in message:
            raise StudyError("学习数据库正在被另一个任务更新，请稍后重试。") from exc
        raise


def connect_db(home: Path, require: bool = True) -> sqlite3.Connection:
    ensure_dirs(home)
    path = db_path(home)
    if require and not path.exists():
        raise StudyError("学习数据库不存在，请先运行 init。")
    connection = sqlite3.connect(path, timeout=SQLITE_BUSY_TIMEOUT_MS / 1000)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    connection.execute("PRAGMA journal_mode = WAL")
    connection.execute("PRAGMA synchronous = FULL")
    connection.execute(f"PRAGMA busy_timeout = {SQLITE_BUSY_TIMEOUT_MS}")
    if require:
        migrate_schema(connection)
    return connection


SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS meta (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS items (
    id TEXT PRIMARY KEY,
    subject TEXT NOT NULL,
    topic TEXT NOT NULL,
    item_type TEXT NOT NULL,
    prompt TEXT NOT NULL,
    answer TEXT NOT NULL,
    aliases_json TEXT NOT NULL,
    tags_json TEXT NOT NULL,
    source TEXT NOT NULL,
    confidence REAL NOT NULL,
    ocr_confidence REAL,
    content_confidence REAL NOT NULL DEFAULT 0.0,
    content_verified INTEGER NOT NULL DEFAULT 0,
    content_hash TEXT NOT NULL UNIQUE,
    card_json TEXT NOT NULL,
    due_utc TEXT NOT NULL,
    fsrs_state INTEGER NOT NULL,
    stability REAL,
    difficulty REAL,
    step INTEGER,
    last_review_utc TEXT,
    created_at_utc TEXT NOT NULL,
    updated_at_utc TEXT NOT NULL,
    last_seen_at_utc TEXT NOT NULL,
    active INTEGER NOT NULL DEFAULT 1
);

CREATE INDEX IF NOT EXISTS idx_items_due ON items(active, due_utc);
CREATE INDEX IF NOT EXISTS idx_items_subject_topic ON items(subject, topic);

CREATE TABLE IF NOT EXISTS pending_imports (
    id TEXT PRIMARY KEY,
    raw_json TEXT NOT NULL,
    reason TEXT NOT NULL,
    content_hash TEXT NOT NULL UNIQUE,
    source TEXT NOT NULL,
    confidence REAL NOT NULL,
    ocr_confidence REAL,
    content_confidence REAL NOT NULL DEFAULT 0.0,
    content_verified INTEGER NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'pending',
    created_at_utc TEXT NOT NULL,
    resolved_at_utc TEXT
);

CREATE INDEX IF NOT EXISTS idx_pending_status ON pending_imports(status, created_at_utc);

CREATE TABLE IF NOT EXISTS import_batches (
    id TEXT PRIMARY KEY,
    source TEXT NOT NULL,
    started_at_utc TEXT NOT NULL,
    finished_at_utc TEXT,
    imported_count INTEGER NOT NULL DEFAULT 0,
    duplicate_count INTEGER NOT NULL DEFAULT 0,
    pending_count INTEGER NOT NULL DEFAULT 0,
    details_json TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS sessions (
    id TEXT PRIMARY KEY,
    kind TEXT NOT NULL,
    started_at_utc TEXT NOT NULL,
    ended_at_utc TEXT,
    minutes INTEGER NOT NULL,
    status TEXT NOT NULL,
    score_first REAL,
    total_answered INTEGER NOT NULL DEFAULT 0,
    report_md TEXT,
    report_csv TEXT
);

CREATE INDEX IF NOT EXISTS idx_sessions_kind_status ON sessions(kind, status, started_at_utc);

CREATE TABLE IF NOT EXISTS session_items (
    session_id TEXT NOT NULL REFERENCES sessions(id),
    item_id TEXT NOT NULL REFERENCES items(id),
    ordinal INTEGER NOT NULL,
    status TEXT NOT NULL DEFAULT 'pending',
    first_result TEXT,
    first_correct INTEGER,
    first_answer TEXT,
    remediation_attempts INTEGER NOT NULL DEFAULT 0,
    answered_at_utc TEXT,
    PRIMARY KEY (session_id, item_id)
);

CREATE TABLE IF NOT EXISTS reviews (
    id TEXT PRIMARY KEY,
    item_id TEXT NOT NULL REFERENCES items(id),
    session_id TEXT NOT NULL REFERENCES sessions(id),
    reviewed_at_utc TEXT NOT NULL,
    rating TEXT NOT NULL,
    answer_text TEXT NOT NULL,
    first_attempt_correct INTEGER,
    is_remediation INTEGER NOT NULL,
    response_ms INTEGER,
    error_reason TEXT,
    card_before_json TEXT NOT NULL,
    card_after_json TEXT NOT NULL,
    due_after_utc TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_reviews_item ON reviews(item_id, reviewed_at_utc);
CREATE INDEX IF NOT EXISTS idx_reviews_session ON reviews(session_id, reviewed_at_utc);

CREATE TABLE IF NOT EXISTS mistakes (
    id TEXT PRIMARY KEY,
    item_id TEXT NOT NULL REFERENCES items(id),
    session_id TEXT NOT NULL REFERENCES sessions(id),
    first_review_id TEXT NOT NULL REFERENCES reviews(id),
    observed_answer TEXT NOT NULL,
    correct_answer TEXT NOT NULL,
    error_reason TEXT,
    resolved INTEGER NOT NULL DEFAULT 0,
    resolved_at_utc TEXT,
    variant_count INTEGER NOT NULL DEFAULT 0,
    created_at_utc TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_mistakes_resolved ON mistakes(resolved, created_at_utc);
"""


def table_columns(connection: sqlite3.Connection, table: str) -> set[str]:
    return {str(row[1]) for row in connection.execute(f"PRAGMA table_info({table})").fetchall()}


def migrate_schema(connection: sqlite3.Connection) -> None:
    """Upgrade existing databases transactionally without rewriting study history."""
    item_columns = table_columns(connection, "items")
    pending_columns = table_columns(connection, "pending_imports")
    if not item_columns or not pending_columns:
        return
    version_row = connection.execute(
        "SELECT value FROM meta WHERE key = 'schema_version'"
    ).fetchone()
    current_version = str(version_row[0]) if version_row else None
    required_item_columns = {"ocr_confidence", "content_confidence", "content_verified"}
    required_pending_columns = {"ocr_confidence", "content_confidence", "content_verified"}
    if (
        current_version == SCHEMA_VERSION
        and required_item_columns.issubset(item_columns)
        and required_pending_columns.issubset(pending_columns)
    ):
        return
    try:
        begin_immediate(connection)
        # Another process may have completed the migration while this connection waited.
        item_columns = table_columns(connection, "items")
        pending_columns = table_columns(connection, "pending_imports")
        version_row = connection.execute(
            "SELECT value FROM meta WHERE key = 'schema_version'"
        ).fetchone()
        current_version = str(version_row[0]) if version_row else None
        if (
            current_version == SCHEMA_VERSION
            and required_item_columns.issubset(item_columns)
            and required_pending_columns.issubset(pending_columns)
        ):
            connection.rollback()
            return
        if "ocr_confidence" not in item_columns:
            connection.execute("ALTER TABLE items ADD COLUMN ocr_confidence REAL")
        if "content_confidence" not in item_columns:
            connection.execute(
                "ALTER TABLE items ADD COLUMN content_confidence REAL NOT NULL DEFAULT 0.0"
            )
            connection.execute("UPDATE items SET content_confidence = confidence")
        if "content_verified" not in item_columns:
            connection.execute(
                "ALTER TABLE items ADD COLUMN content_verified INTEGER NOT NULL DEFAULT 0"
            )
        if "ocr_confidence" not in pending_columns:
            connection.execute("ALTER TABLE pending_imports ADD COLUMN ocr_confidence REAL")
        if "content_confidence" not in pending_columns:
            connection.execute(
                """
                ALTER TABLE pending_imports
                ADD COLUMN content_confidence REAL NOT NULL DEFAULT 0.0
                """
            )
            connection.execute("UPDATE pending_imports SET content_confidence = confidence")
        if "content_verified" not in pending_columns:
            connection.execute(
                """
                ALTER TABLE pending_imports
                ADD COLUMN content_verified INTEGER NOT NULL DEFAULT 0
                """
            )
        connection.execute(
            "INSERT OR REPLACE INTO meta(key, value) VALUES('schema_version', ?)",
            (SCHEMA_VERSION,),
        )
        connection.commit()
    except Exception:
        connection.rollback()
        raise


def initialize(home: Path) -> dict[str, Any]:
    ensure_dirs(home)
    config = load_config(home, create=True)
    connection = connect_db(home, require=False)
    try:
        connection.executescript(SCHEMA_SQL)
        migrate_schema(connection)
        begin_immediate(connection)
        now = iso_utc()
        connection.execute(
            "INSERT OR REPLACE INTO meta(key, value) VALUES('schema_version', ?)",
            (SCHEMA_VERSION,),
        )
        connection.execute(
            "INSERT OR IGNORE INTO meta(key, value) VALUES('created_at_utc', ?)",
            (now,),
        )
        connection.commit()
        dashboard = generate_dashboard(connection, home, config, write=True)
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    backup = backup_database(home, "initialization", config)
    return {
        "initialized": True,
        "home": str(home),
        "database": str(db_path(home)),
        "dashboard": dashboard["report"],
        "backup": str(backup),
    }


def scheduler_from_config(config: dict[str, Any]) -> Scheduler:
    return Scheduler(
        desired_retention=float(config["desired_retention"]),
        learning_steps=(timedelta(minutes=1), timedelta(minutes=10)),
        relearning_steps=(timedelta(minutes=10),),
        maximum_interval=36500,
        enable_fuzzing=bool(config.get("enable_fuzzing", True)),
    )


def random_card(now: datetime | None = None) -> Card:
    card_id = int.from_bytes(uuid.uuid4().bytes[:8], "big") & ((1 << 63) - 1)
    return Card(card_id=card_id, due=now or utc_now())


def card_fields(card: Card) -> dict[str, Any]:
    data = card.to_dict()
    return {
        "card_json": json.dumps(data, ensure_ascii=False, sort_keys=True),
        "due_utc": data["due"],
        "fsrs_state": int(data["state"]),
        "stability": data["stability"],
        "difficulty": data["difficulty"],
        "step": data["step"],
        "last_review_utc": data["last_review"],
    }


def backup_hash_path(path: Path) -> Path:
    return path.with_name(path.name + ".sha256")


def prune_backups(home: Path, config: dict[str, Any], protected: Path) -> dict[str, int]:
    count_limit = max(1, int(config.get("backup_retention_count", 30)))
    byte_limit = max(
        1,
        int(float(config.get("backup_max_total_mb", 2048)) * 1024 * 1024),
    )
    backups = sorted(
        (home / "backups").glob("*.sqlite3"),
        key=lambda path: (path.stat().st_mtime_ns, path.name),
    )
    total_bytes = sum(path.stat().st_size for path in backups)
    removed = 0
    removed_bytes = 0
    for candidate in list(backups):
        if len(backups) <= count_limit and total_bytes <= byte_limit:
            break
        if candidate == protected:
            continue
        size = candidate.stat().st_size
        candidate.unlink()
        backup_hash_path(candidate).unlink(missing_ok=True)
        backups.remove(candidate)
        total_bytes -= size
        removed += 1
        removed_bytes += size
    return {"removed": removed, "removed_bytes": removed_bytes}


def verify_backup(path: Path) -> dict[str, Any]:
    connection = sqlite3.connect(f"{path.as_uri()}?mode=ro", uri=True)
    try:
        sqlite_integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
    finally:
        connection.close()
    hash_path = backup_hash_path(path)
    expected = (
        hash_path.read_text(encoding="utf-8").split(maxsplit=1)[0] if hash_path.is_file() else None
    )
    actual = file_sha256(path)
    return {
        "sqlite_integrity": sqlite_integrity,
        "sha256_manifest": str(hash_path) if hash_path.is_file() else None,
        "sha256_matches": expected is not None and expected.casefold() == actual.casefold(),
    }


def backup_database(home: Path, reason: str, config: dict[str, Any] | None = None) -> Path:
    ensure_dirs(home)
    active_config = config or load_config(home)
    source_path = db_path(home)
    if not source_path.exists():
        raise StudyError("学习数据库不存在，无法备份。")
    stamp = utc_now().strftime("%Y%m%dT%H%M%S.%fZ")
    safe_reason = re.sub(r"[^a-zA-Z0-9_-]+", "-", reason).strip("-") or "manual"
    target = home / "backups" / f"study-{stamp}-{safe_reason}.sqlite3"
    source = sqlite3.connect(source_path)
    destination = sqlite3.connect(target)
    try:
        source.execute("PRAGMA wal_checkpoint(FULL)")
        source.backup(destination)
        integrity = destination.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise StudyError(f"备份完整性检查失败：{integrity}")
    except Exception:
        destination.close()
        source.close()
        target.unlink(missing_ok=True)
        raise
    finally:
        try:
            destination.close()
        finally:
            source.close()
    digest = file_sha256(target)
    atomic_write_text(backup_hash_path(target), f"{digest}  {target.name}\n")
    prune_backups(home, active_config, target)
    return target


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def split_multi(value: Any) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, (list, tuple, set)):
        return [clean_text(item) for item in value if clean_text(item)]
    return [part.strip() for part in re.split(r"[|;；\n]+", str(value)) if part.strip()]


def canonicalize_mapping(raw: dict[str, Any]) -> dict[str, Any]:
    canonical: dict[str, Any] = {}
    for key, value in raw.items():
        normalized_key = clean_text(key).lower()
        canonical[HEADER_ALIASES.get(normalized_key, normalized_key)] = value
    return canonical


def bounded_confidence(value: Any, default: float) -> float:
    try:
        confidence = float(default if value is None else value)
    except (TypeError, ValueError):
        confidence = 0.0
    return min(1.0, max(0.0, confidence))


def parse_boolean(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return clean_text(value).casefold() in {"1", "true", "yes", "y", "是", "已确认"}


def is_ocr_source(source: str, data: dict[str, Any]) -> bool:
    if "ocr_confidence" in data:
        return True
    normalized = source.casefold()
    markers = ("ocr", "vision", "image", "scan", "图片", "扫描")
    return any(marker in normalized for marker in markers)


def normalize_entry(
    raw: dict[str, Any],
    default_subject: str,
    default_topic: str,
    default_source: str,
    default_confidence: float,
) -> dict[str, Any]:
    data = canonicalize_mapping(raw)
    item_type = clean_text(data.get("type") or "term").lower()
    item_type = TYPE_ALIASES.get(item_type, item_type)
    if item_type not in ITEM_TYPES:
        item_type = "qa"
    source = clean_text(data.get("source") or default_source or "manual")
    legacy_confidence = bounded_confidence(data.get("confidence"), default_confidence)
    ocr_derived = is_ocr_source(source, data)
    content_verified = parse_boolean(data.get("content_verified"))
    if data.get("ocr_confidence") is not None:
        ocr_confidence: float | None = bounded_confidence(data.get("ocr_confidence"), 0.0)
    elif ocr_derived:
        ocr_confidence = legacy_confidence
    else:
        ocr_confidence = None
    if data.get("content_confidence") is not None:
        content_confidence = bounded_confidence(data.get("content_confidence"), 0.0)
    elif ocr_derived:
        content_confidence = 1.0 if content_verified else 0.0
    else:
        content_confidence = legacy_confidence
    if source.casefold() == "model-assisted":
        content_confidence = min(content_confidence, 0.85)
    return {
        "subject": clean_text(data.get("subject") or default_subject or "英语"),
        "topic": clean_text(data.get("topic") or default_topic or "未分类"),
        "type": item_type,
        "prompt": clean_text(data.get("prompt")),
        "answer": clean_text(data.get("answer")),
        "aliases": split_multi(data.get("aliases")),
        "tags": split_multi(data.get("tags")),
        "source": source,
        "confidence": content_confidence,
        "ocr_confidence": ocr_confidence,
        "content_confidence": content_confidence,
        "content_verified": content_verified,
    }


def content_hash(entry: dict[str, Any], include_answer: bool = True) -> str:
    parts = [
        entry.get("subject", "").casefold(),
        entry.get("topic", "").casefold(),
        entry.get("type", "").casefold(),
        entry.get("prompt", "").casefold(),
    ]
    if include_answer:
        parts.append(entry.get("answer", "").casefold())
    payload = "\x1f".join(parts)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def read_text_with_fallback(path: Path) -> str:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise StudyError(f"无法识别文本编码：{path.name}")


def row_mapping(headers: Sequence[Any], row: Sequence[Any]) -> dict[str, Any]:
    return {clean_text(key): value for key, value in zip(headers, row, strict=False)}


def looks_like_header(row: Sequence[Any]) -> bool:
    mapped = {
        HEADER_ALIASES.get(clean_text(value).lower(), clean_text(value).lower())
        for value in row
        if value is not None
    }
    return "prompt" in mapped or "answer" in mapped


def parse_delimited(path: Path) -> list[dict[str, Any]]:
    text = read_text_with_fallback(path)
    sample = text[:4096]
    if path.suffix.lower() == ".tsv":
        delimiter = "\t"
    else:
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",\t").delimiter
        except csv.Error:
            delimiter = ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    rows = [row for row in rows if any(clean_text(cell) for cell in row)]
    if not rows:
        return []
    if looks_like_header(rows[0]):
        headers = rows[0]
        return [row_mapping(headers, row) for row in rows[1:]]
    return [
        {"prompt": row[0] if row else "", "answer": row[1] if len(row) > 1 else ""} for row in rows
    ]


def parse_excel(path: Path) -> list[dict[str, Any]]:
    with zipfile.ZipFile(path) as archive:
        expanded = sum(info.file_size for info in archive.infolist())
        if expanded > MAX_EXPANDED_DOCUMENT_BYTES:
            raise StudyError("Excel 解压后内容超过 200 MB 安全上限。")
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    output: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            rows = [
                list(row)
                for row in sheet.iter_rows(values_only=True)
                if any(clean_text(cell) for cell in row)
            ]
            if not rows:
                continue
            if looks_like_header(rows[0]):
                headers = rows[0]
                for row in rows[1:]:
                    mapped = row_mapping(headers, row)
                    mapped.setdefault("source", f"{path.name}#{sheet.title}")
                    output.append(mapped)
            else:
                for row in rows:
                    output.append(
                        {
                            "prompt": row[0] if row else "",
                            "answer": row[1] if len(row) > 1 else "",
                            "source": f"{path.name}#{sheet.title}",
                        }
                    )
    finally:
        workbook.close()
    return output


def parse_text_lines(text: str, source: str) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for raw_line in text.splitlines():
        line = re.sub(r"^\s*(?:[-*•]|\d+[.)、])\s*", "", raw_line).strip()
        line = line.lstrip("#").strip()
        if not line:
            continue
        parts: list[str] | None = None
        for delimiter in ("\t", "::", "=>"):
            if delimiter in line:
                parts = line.split(delimiter, 1)
                break
        if parts is None:
            colon = re.match(r"^([^:：]{1,100})\s*[:：]\s*(.+)$", line)
            if colon:
                parts = [colon.group(1), colon.group(2)]
        if parts is None:
            spaced = re.split(r"\s{2,}", line, maxsplit=1)
            if len(spaced) == 2:
                parts = spaced
        if parts is None:
            output.append({"prompt": line, "answer": "", "source": source})
        else:
            output.append(
                {"prompt": clean_text(parts[0]), "answer": clean_text(parts[1]), "source": source}
            )
    return output


def parse_pdf(path: Path) -> list[dict[str, Any]]:
    reader = PdfReader(path)
    if reader.is_encrypted and reader.decrypt("") == 0:
        raise StudyError("PDF 已加密，无法读取。")
    if len(reader.pages) > MAX_PDF_PAGES:
        raise StudyError(f"PDF 页数超过安全上限 {MAX_PDF_PAGES}。")
    pages: list[str] = []
    extracted_chars = 0
    for page in reader.pages:
        extracted = page.extract_text() or ""
        if extracted.strip():
            extracted_chars += len(extracted)
            if extracted_chars > MAX_EXTRACTED_TEXT_CHARS:
                raise StudyError("PDF 提取文字超过 500 万字符安全上限。")
            pages.append(f"\n{extracted}\n")
    if not pages:
        raise StudyError("PDF 没有可提取文字；请让技能按扫描图片识别并生成 JSON。")
    return parse_text_lines("\n".join(pages), path.name)


def load_entries(path: Path) -> list[dict[str, Any]]:
    if not path.exists() or not path.is_file():
        raise StudyError(f"导入文件不存在：{path}")
    if path.stat().st_size > MAX_IMPORT_BYTES:
        raise StudyError(f"导入文件超过 {MAX_IMPORT_BYTES // (1024 * 1024)} MB 安全上限。")
    suffix = path.suffix.lower()
    if suffix in {".json", ".jsonl"}:
        text = read_text_with_fallback(path)
        if suffix == ".jsonl":
            return [json.loads(line) for line in text.splitlines() if line.strip()]
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            parsed = parsed.get("items", [parsed])
        if not isinstance(parsed, list) or not all(isinstance(item, dict) for item in parsed):
            raise StudyError("JSON 必须是对象数组或包含 items 数组。")
        return parsed
    if suffix in {".csv", ".tsv"}:
        return parse_delimited(path)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_excel(path)
    if suffix == ".pdf":
        return parse_pdf(path)
    if suffix in {".txt", ".md"}:
        return parse_text_lines(read_text_with_fallback(path), path.name)
    raise StudyError(f"暂不支持该文件类型：{suffix}")


def insert_pending(connection: sqlite3.Connection, entry: dict[str, Any], reason: str) -> bool:
    raw_json = json.dumps(entry, ensure_ascii=False, sort_keys=True)
    pending_hash = hashlib.sha256(raw_json.encode("utf-8")).hexdigest()
    result = connection.execute(
        """
        INSERT OR IGNORE INTO pending_imports(
            id, raw_json, reason, content_hash, source, confidence,
            ocr_confidence, content_confidence, content_verified,
            status, created_at_utc
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?)
        """,
        (
            uuid.uuid4().hex,
            raw_json,
            reason,
            pending_hash,
            entry["source"],
            entry["confidence"],
            entry["ocr_confidence"],
            entry["content_confidence"],
            int(entry["content_verified"]),
            iso_utc(),
        ),
    )
    return result.rowcount == 1


def insert_item(connection: sqlite3.Connection, entry: dict[str, Any]) -> bool:
    now = utc_now()
    card = random_card(now)
    fields = card_fields(card)
    result = connection.execute(
        """
        INSERT OR IGNORE INTO items(
            id, subject, topic, item_type, prompt, answer, aliases_json, tags_json,
            source, confidence, ocr_confidence, content_confidence, content_verified,
            content_hash, card_json, due_utc, fsrs_state,
            stability, difficulty, step, last_review_utc, created_at_utc,
            updated_at_utc, last_seen_at_utc, active
        ) VALUES (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1
        )
        """,
        (
            uuid.uuid4().hex,
            entry["subject"],
            entry["topic"],
            entry["type"],
            entry["prompt"],
            entry["answer"],
            json.dumps(entry["aliases"], ensure_ascii=False),
            json.dumps(entry["tags"], ensure_ascii=False),
            entry["source"],
            entry["confidence"],
            entry["ocr_confidence"],
            entry["content_confidence"],
            int(entry["content_verified"]),
            content_hash(entry),
            fields["card_json"],
            fields["due_utc"],
            fields["fsrs_state"],
            fields["stability"],
            fields["difficulty"],
            fields["step"],
            fields["last_review_utc"],
            iso_utc(now),
            iso_utc(now),
            iso_utc(now),
        ),
    )
    if result.rowcount == 0:
        connection.execute(
            "UPDATE items SET last_seen_at_utc = ? WHERE content_hash = ?",
            (iso_utc(now), content_hash(entry)),
        )
        return False
    return True


def import_file(
    home: Path,
    config: dict[str, Any],
    path: Path,
    default_subject: str,
    default_topic: str,
    default_confidence: float,
) -> dict[str, Any]:
    path = resolve_workspace_input(path)
    raw_entries = load_entries(path)
    batch_id = uuid.uuid4().hex
    connection = connect_db(home)
    imported = duplicates = pending = pending_duplicates = 0
    try:
        begin_immediate(connection)
        connection.execute(
            """
            INSERT INTO import_batches(id, source, started_at_utc, details_json)
            VALUES (?, ?, ?, ?)
            """,
            (
                batch_id,
                str(path),
                iso_utc(),
                json.dumps({"raw_count": len(raw_entries)}, ensure_ascii=False),
            ),
        )
        threshold = float(config["auto_import_confidence"])
        for raw in raw_entries:
            entry = normalize_entry(
                raw,
                default_subject,
                default_topic,
                path.name,
                default_confidence,
            )
            reasons: list[str] = []
            if not entry["prompt"]:
                reasons.append("缺少题面")
            if not entry["answer"]:
                reasons.append("缺少答案")
            if entry["ocr_confidence"] is not None and not entry["content_verified"]:
                reasons.append("OCR 内容尚未人工确认")
            if entry["content_confidence"] < threshold:
                reasons.append(f"内容置信度低于 {threshold:.2f}")
            if reasons:
                if insert_pending(connection, entry, "；".join(reasons)):
                    pending += 1
                else:
                    pending_duplicates += 1
                continue
            if insert_item(connection, entry):
                imported += 1
            else:
                duplicates += 1
        connection.execute(
            """
            UPDATE import_batches
            SET finished_at_utc = ?, imported_count = ?, duplicate_count = ?,
                pending_count = ?, details_json = ?
            WHERE id = ?
            """,
            (
                iso_utc(),
                imported,
                duplicates,
                pending,
                json.dumps(
                    {
                        "raw_count": len(raw_entries),
                        "pending_duplicates": pending_duplicates,
                    },
                    ensure_ascii=False,
                ),
                batch_id,
            ),
        )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    backup = backup_database(home, "import", config)
    return {
        "batch_id": batch_id,
        "source": str(path),
        "read": len(raw_entries),
        "imported": imported,
        "duplicates": duplicates,
        "pending": pending,
        "pending_duplicates": pending_duplicates,
        "backup": str(backup),
    }


def pending_list(home: Path) -> dict[str, Any]:
    connection = connect_db(home)
    try:
        rows = connection.execute(
            """
            SELECT id, raw_json, reason, source, confidence, created_at_utc
            FROM pending_imports WHERE status = 'pending'
            ORDER BY created_at_utc
            """
        ).fetchall()
        items = []
        for row in rows:
            raw = json.loads(row["raw_json"])
            items.append(
                {
                    "pending_id": row["id"],
                    "reason": row["reason"],
                    "source": row["source"],
                    "confidence": row["confidence"],
                    "created_local": local_display(row["created_at_utc"]),
                    "entry": raw,
                }
            )
        return {"count": len(items), "items": items}
    finally:
        connection.close()


def pending_resolve(home: Path, config: dict[str, Any], args: argparse.Namespace) -> dict[str, Any]:
    connection = connect_db(home)
    try:
        begin_immediate(connection)
        row = connection.execute(
            "SELECT * FROM pending_imports WHERE id = ? AND status = 'pending'",
            (args.pending_id,),
        ).fetchone()
        if row is None:
            raise StudyError("待确认条目不存在或已处理。")
        raw = json.loads(row["raw_json"])
        for key in ("subject", "topic", "type", "prompt", "answer"):
            value = getattr(args, key, None)
            if value is not None:
                raw[key] = value
        raw["confidence"] = 1.0
        raw["content_confidence"] = 1.0
        raw["content_verified"] = True
        entry = normalize_entry(raw, "英语", "未分类", row["source"], 1.0)
        if not entry["prompt"] or not entry["answer"]:
            raise StudyError("确认条目必须同时有题面和答案。")
        created = insert_item(connection, entry)
        connection.execute(
            """
            UPDATE pending_imports SET status = 'resolved', resolved_at_utc = ?
            WHERE id = ?
            """,
            (iso_utc(), args.pending_id),
        )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    backup = backup_database(home, "pending-resolve", config)
    return {
        "pending_id": args.pending_id,
        "created": created,
        "duplicate": not created,
        "backup": str(backup),
    }


def safe_retrievability(scheduler: Scheduler, card: Card, now: datetime) -> float:
    try:
        return float(scheduler.get_card_retrievability(card, now))
    except (TypeError, ValueError, ZeroDivisionError):
        return 0.0


def item_payload(row: sqlite3.Row, scheduler: Scheduler, now: datetime) -> dict[str, Any]:
    card = Card.from_dict(json.loads(row["card_json"]))
    return {
        "item_id": row["id"],
        "subject": row["subject"],
        "topic": row["topic"],
        "type": row["item_type"],
        "prompt": row["prompt"],
        "answer": row["answer"],
        "aliases": json.loads(row["aliases_json"]),
        "tags": json.loads(row["tags_json"]),
        "due_utc": row["due_utc"],
        "due_local": local_display(row["due_utc"]),
        "retrievability": round(safe_retrievability(scheduler, card, now), 4),
    }


def daily_candidates(
    connection: sqlite3.Connection,
    scheduler: Scheduler,
    now: datetime,
    limit: int,
) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT * FROM items
        WHERE active = 1 AND due_utc <= ?
        ORDER BY due_utc ASC
        """,
        (iso_utc(end_of_local_day_utc(now)),),
    ).fetchall()
    payloads = [item_payload(row, scheduler, now) for row in rows]

    def priority(item: dict[str, Any]) -> tuple[Any, ...]:
        due = parse_dt(item["due_utc"]) or now
        overdue_group = 0 if due <= now else 1
        return (overdue_group, item["retrievability"], due)

    return sorted(payloads, key=priority)[:limit]


def test_candidates(
    connection: sqlite3.Connection,
    scheduler: Scheduler,
    now: datetime,
    limit: int,
) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT i.*,
               COUNT(r.id) AS review_count,
               COALESCE(SUM(CASE
                   WHEN r.is_remediation = 0 AND r.first_attempt_correct = 0 THEN 1
                   ELSE 0 END), 0) AS wrong_count,
               AVG(CASE
                   WHEN r.is_remediation = 0 THEN r.first_attempt_correct
                   ELSE NULL END) AS accuracy,
               MAX(r.reviewed_at_utc) AS last_reviewed
        FROM items i
        LEFT JOIN reviews r ON r.item_id = i.id
        WHERE i.active = 1
        GROUP BY i.id
        """
    ).fetchall()
    ranked: list[tuple[float, dict[str, Any]]] = []
    for row in rows:
        payload = item_payload(row, scheduler, now)
        accuracy = float(row["accuracy"]) if row["accuracy"] is not None else 0.5
        wrong_count = int(row["wrong_count"])
        last_review = parse_dt(row["last_reviewed"])
        age_days = 30.0 if last_review is None else min(90.0, (now - last_review).days)
        weakness = (
            wrong_count * 3.0
            + (1.0 - accuracy) * 2.0
            + (1.0 - payload["retrievability"])
            + age_days / 90.0
        )
        payload["history"] = {
            "review_count": int(row["review_count"]),
            "wrong_count": wrong_count,
            "first_accuracy": round(accuracy, 4),
        }
        ranked.append((weakness, payload))

    ranked.sort(key=lambda pair: (-pair[0], pair[1]["subject"], pair[1]["topic"]))
    buckets: dict[str, deque[dict[str, Any]]] = defaultdict(deque)
    for _, payload in ranked:
        buckets[payload["subject"]].append(payload)

    output: list[dict[str, Any]] = []
    subjects = deque(sorted(buckets, key=lambda name: (-len(buckets[name]), name)))
    while subjects and len(output) < limit:
        subject = subjects.popleft()
        bucket = buckets[subject]
        if bucket:
            output.append(bucket.popleft())
        if bucket:
            subjects.append(subject)
    return output


def session_existing(
    connection: sqlite3.Connection, kind: str, now: datetime
) -> sqlite3.Row | None:
    return connection.execute(
        """
        SELECT * FROM sessions
        WHERE kind = ? AND status = 'active'
        ORDER BY started_at_utc DESC LIMIT 1
        """,
        (kind,),
    ).fetchone()


def session_items_payload(
    connection: sqlite3.Connection,
    session_id: str,
    scheduler: Scheduler,
    now: datetime,
) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT i.*, si.ordinal, si.status AS session_status, si.first_result,
               si.first_correct, si.first_answer, si.remediation_attempts
        FROM session_items si
        JOIN items i ON i.id = si.item_id
        WHERE si.session_id = ?
        ORDER BY si.ordinal
        """,
        (session_id,),
    ).fetchall()
    output = []
    for row in rows:
        payload = item_payload(row, scheduler, now)
        payload.update(
            {
                "ordinal": row["ordinal"],
                "session_status": row["session_status"],
                "first_result": row["first_result"],
                "first_correct": row["first_correct"],
                "first_answer": row["first_answer"],
                "remediation_attempts": row["remediation_attempts"],
            }
        )
        output.append(payload)
    return output


def start_session(
    home: Path,
    config: dict[str, Any],
    kind: str,
    minutes: int | None,
    limit: int | None,
    force_new: bool,
) -> dict[str, Any]:
    if kind not in {"daily", "weekly", "monthly"}:
        raise StudyError("会话类型必须是 daily、weekly 或 monthly。")
    defaults = {
        "daily": (
            int(config["daily_minutes"]),
            int(config["daily_question_limit"]),
        ),
        "weekly": (
            int(config["weekly_minutes"]),
            int(config["weekly_questions"]),
        ),
        "monthly": (
            int(config["monthly_minutes"]),
            int(config["monthly_questions"]),
        ),
    }
    session_minutes = int(minutes or defaults[kind][0])
    item_limit = int(limit or defaults[kind][1])
    if session_minutes <= 0 or item_limit <= 0:
        raise StudyError("时长和题目数量必须是正整数。")

    connection = connect_db(home)
    scheduler = scheduler_from_config(config)
    now = utc_now()
    try:
        begin_immediate(connection)
        if not force_new and (existing := session_existing(connection, kind, now)):
            items = session_items_payload(connection, existing["id"], scheduler, now)
            if items:
                return {
                    "session_id": existing["id"],
                    "kind": kind,
                    "minutes": existing["minutes"],
                    "reused": True,
                    "count": len(items),
                    "items": items,
                }
            connection.execute(
                """
                UPDATE sessions SET status = 'empty', ended_at_utc = ?
                WHERE id = ?
                """,
                (iso_utc(now), existing["id"]),
            )

        if kind == "daily":
            selected = daily_candidates(connection, scheduler, now, item_limit)
        else:
            selected = test_candidates(connection, scheduler, now, item_limit)
        session_id = uuid.uuid4().hex
        connection.execute(
            """
            INSERT INTO sessions(id, kind, started_at_utc, minutes, status)
            VALUES (?, ?, ?, ?, 'active')
            """,
            (session_id, kind, iso_utc(now), session_minutes),
        )
        for ordinal, item in enumerate(selected, start=1):
            connection.execute(
                """
                INSERT INTO session_items(session_id, item_id, ordinal)
                VALUES (?, ?, ?)
                """,
                (session_id, item["item_id"], ordinal),
            )
        connection.commit()
        return {
            "session_id": session_id,
            "kind": kind,
            "minutes": session_minutes,
            "reused": False,
            "count": len(selected),
            "items": selected,
            "instruction": "一次只展示一道题；答案字段仅供判分，不要提前告诉学习者。",
        }
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def show_session(home: Path, config: dict[str, Any], session_id: str) -> dict[str, Any]:
    connection = connect_db(home)
    scheduler = scheduler_from_config(config)
    now = utc_now()
    try:
        session = connection.execute(
            "SELECT * FROM sessions WHERE id = ?", (session_id,)
        ).fetchone()
        if session is None:
            raise StudyError("学习会话不存在。")
        items = session_items_payload(connection, session_id, scheduler, now)
        return {
            "session_id": session_id,
            "kind": session["kind"],
            "status": session["status"],
            "minutes": session["minutes"],
            "started_local": local_display(session["started_at_utc"]),
            "count": len(items),
            "items": items,
        }
    finally:
        connection.close()


RESULT_ALIASES = {
    "wrong": "again",
    "incorrect": "again",
    "correct": "good",
    "again": "again",
    "hard": "hard",
    "good": "good",
    "easy": "easy",
}

if DEPENDENCY_IMPORT_ERROR is None:
    RATING_MAP = {
        "again": Rating.Again,
        "hard": Rating.Hard,
        "good": Rating.Good,
        "easy": Rating.Easy,
    }
else:
    RATING_MAP = {}


def remediation_limits(config: dict[str, Any]) -> tuple[int, int]:
    per_item = max(1, int(config.get("max_remediation_attempts_per_item", 3)))
    per_session = max(1, int(config.get("max_remediation_attempts_per_session", 12)))
    return per_item, per_session


def session_remediation_attempts(connection: sqlite3.Connection, session_id: str) -> int:
    return int(
        connection.execute(
            """
            SELECT COALESCE(SUM(remediation_attempts), 0)
            FROM session_items WHERE session_id = ?
            """,
            (session_id,),
        ).fetchone()[0]
    )


def record_answer(
    home: Path,
    config: dict[str, Any],
    session_id: str,
    item_id: str,
    result: str,
    answer_text: str,
    reason: str | None,
    response_ms: int | None,
    remediation: bool,
) -> dict[str, Any]:
    normalized_result = RESULT_ALIASES.get(result.lower())
    if normalized_result is None:
        raise StudyError("result 必须是 again、hard、good 或 easy。")
    now = utc_now()
    connection = connect_db(home)
    try:
        begin_immediate(connection)
        row = connection.execute(
            """
            SELECT s.status AS session_status, s.kind, si.status AS item_status,
                   si.first_result, si.remediation_attempts, i.*
            FROM session_items si
            JOIN sessions s ON s.id = si.session_id
            JOIN items i ON i.id = si.item_id
            WHERE si.session_id = ? AND si.item_id = ?
            """,
            (session_id, item_id),
        ).fetchone()
        if row is None:
            raise StudyError("该题不属于此学习会话。")
        if row["session_status"] != "active":
            raise StudyError("学习会话已经结束。")
        before_json = row["card_json"]
        review_id = uuid.uuid4().hex
        per_item_limit, per_session_limit = remediation_limits(config)

        if remediation:
            if row["first_result"] != "again":
                raise StudyError("只有首次答错的题目才能记录补救变式。")
            if row["item_status"] == "done":
                raise StudyError("该题已经完成补救。")
            if row["item_status"] == "remediation_exhausted":
                raise StudyError("该题补救次数已经达到硬上限，请继续下一题。")
            current_item_attempts = int(row["remediation_attempts"])
            current_session_attempts = session_remediation_attempts(connection, session_id)
            if current_item_attempts >= per_item_limit:
                raise StudyError("该知识点补救次数已经达到硬上限。")
            if current_session_attempts >= per_session_limit:
                raise StudyError("本场补救总次数已经达到硬上限。")
            correct = normalized_result != "again"
            next_item_attempts = current_item_attempts + 1
            next_session_attempts = current_session_attempts + 1
            item_exhausted = not correct and next_item_attempts >= per_item_limit
            session_exhausted = not correct and next_session_attempts >= per_session_limit
            exhausted = item_exhausted or session_exhausted
            new_status = (
                "done" if correct else "remediation_exhausted" if exhausted else "remediation"
            )
            connection.execute(
                """
                INSERT INTO reviews(
                    id, item_id, session_id, reviewed_at_utc, rating, answer_text,
                    first_attempt_correct, is_remediation, response_ms, error_reason,
                    card_before_json, card_after_json, due_after_utc
                ) VALUES (?, ?, ?, ?, ?, ?, NULL, 1, ?, ?, ?, ?, ?)
                """,
                (
                    review_id,
                    item_id,
                    session_id,
                    iso_utc(now),
                    "RemediationCorrect" if correct else "RemediationAgain",
                    answer_text,
                    response_ms,
                    reason,
                    before_json,
                    before_json,
                    row["due_utc"],
                ),
            )
            connection.execute(
                """
                UPDATE session_items
                SET status = ?, remediation_attempts = remediation_attempts + 1,
                    answered_at_utc = ?
                WHERE session_id = ? AND item_id = ?
                """,
                (new_status, iso_utc(now), session_id, item_id),
            )
            if session_exhausted:
                connection.execute(
                    """
                    UPDATE session_items SET status = 'remediation_exhausted'
                    WHERE session_id = ? AND status = 'remediation'
                    """,
                    (session_id,),
                )
            if correct:
                connection.execute(
                    """
                    UPDATE mistakes
                    SET resolved = 1, resolved_at_utc = ?,
                        variant_count = variant_count + 1
                    WHERE session_id = ? AND item_id = ? AND resolved = 0
                    """,
                    (iso_utc(now), session_id, item_id),
                )
            else:
                connection.execute(
                    """
                    UPDATE mistakes
                    SET variant_count = variant_count + 1,
                        error_reason = COALESCE(?, error_reason)
                    WHERE session_id = ? AND item_id = ? AND resolved = 0
                    """,
                    (reason, session_id, item_id),
                )
            connection.commit()
            return {
                "session_id": session_id,
                "item_id": item_id,
                "remediation": True,
                "correct": correct,
                "needs_another_variant": not correct and not exhausted,
                "remediation_exhausted": exhausted,
                "exhaustion_reason": (
                    "item_limit"
                    if item_exhausted
                    else "session_limit"
                    if session_exhausted
                    else None
                ),
                "item_remediation_attempts": next_item_attempts,
                "item_remediation_limit": per_item_limit,
                "session_remediation_attempts": next_session_attempts,
                "session_remediation_limit": per_session_limit,
                "first_fsrs_rating_preserved": "Again",
                "due_local": local_display(row["due_utc"]),
                "answer": row["answer"],
            }

        if row["first_result"] is not None:
            raise StudyError("首次作答已经记录；错题变式必须添加 --remediation。")
        rating = RATING_MAP[normalized_result]
        first_correct = normalized_result != "again"
        card = Card.from_dict(json.loads(before_json))
        scheduler = scheduler_from_config(config)
        reviewed_card, _ = scheduler.review_card(
            card,
            rating,
            review_datetime=now,
            review_duration=response_ms,
        )
        fields = card_fields(reviewed_card)
        connection.execute(
            """
            UPDATE items
            SET card_json = ?, due_utc = ?, fsrs_state = ?, stability = ?,
                difficulty = ?, step = ?, last_review_utc = ?, updated_at_utc = ?
            WHERE id = ?
            """,
            (
                fields["card_json"],
                fields["due_utc"],
                fields["fsrs_state"],
                fields["stability"],
                fields["difficulty"],
                fields["step"],
                fields["last_review_utc"],
                iso_utc(now),
                item_id,
            ),
        )
        connection.execute(
            """
            INSERT INTO reviews(
                id, item_id, session_id, reviewed_at_utc, rating, answer_text,
                first_attempt_correct, is_remediation, response_ms, error_reason,
                card_before_json, card_after_json, due_after_utc
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?)
            """,
            (
                review_id,
                item_id,
                session_id,
                iso_utc(now),
                normalized_result.capitalize(),
                answer_text,
                int(first_correct),
                response_ms,
                reason,
                before_json,
                fields["card_json"],
                fields["due_utc"],
            ),
        )
        existing_session_attempts = session_remediation_attempts(connection, session_id)
        remediation_available = existing_session_attempts < per_session_limit
        item_status = (
            "done"
            if first_correct
            else "remediation"
            if remediation_available
            else "remediation_exhausted"
        )
        connection.execute(
            """
            UPDATE session_items
            SET status = ?, first_result = ?, first_correct = ?, first_answer = ?,
                answered_at_utc = ?
            WHERE session_id = ? AND item_id = ?
            """,
            (
                item_status,
                normalized_result,
                int(first_correct),
                answer_text,
                iso_utc(now),
                session_id,
                item_id,
            ),
        )
        if not first_correct and not remediation_available:
            connection.execute(
                """
                UPDATE session_items SET status = 'remediation_exhausted'
                WHERE session_id = ? AND status = 'remediation'
                """,
                (session_id,),
            )
        if not first_correct:
            connection.execute(
                """
                INSERT INTO mistakes(
                    id, item_id, session_id, first_review_id, observed_answer,
                    correct_answer, error_reason, created_at_utc
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    uuid.uuid4().hex,
                    item_id,
                    session_id,
                    review_id,
                    answer_text,
                    row["answer"],
                    reason,
                    iso_utc(now),
                ),
            )
        connection.commit()
        return {
            "session_id": session_id,
            "item_id": item_id,
            "remediation": False,
            "first_result": normalized_result,
            "first_correct": first_correct,
            "needs_remediation": not first_correct and remediation_available,
            "remediation_exhausted": not first_correct and not remediation_available,
            "exhaustion_reason": (
                "session_limit" if not first_correct and not remediation_available else None
            ),
            "item_remediation_limit": per_item_limit,
            "session_remediation_limit": per_session_limit,
            "answer": row["answer"],
            "aliases": json.loads(row["aliases_json"]),
            "next_due_utc": fields["due_utc"],
            "next_due_local": local_display(fields["due_utc"]),
        }
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


MARKDOWN_CONTROL_PATTERN = re.compile(r"([\\`*_{}\[\]()#+\-.!|])")


def markdown_escape(value: Any) -> str:
    """Render untrusted study text as table/list data, not active Markdown or HTML."""
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[^\S\n]+", " ", text).strip()
    escaped_html = html.escape(text, quote=False)
    escaped_markdown = MARKDOWN_CONTROL_PATTERN.sub(r"\\\1", escaped_html)
    return escaped_markdown.replace("\n", "<br>")


CSV_FORMULA_PREFIXES = ("=", "+", "-", "@")


def csv_safe_cell(value: Any) -> str:
    """Prevent spreadsheet formula execution while preserving displayed content."""
    text = "" if value is None else str(value)
    if text.lstrip().startswith(CSV_FORMULA_PREFIXES):
        return "'" + text
    return text


def write_csv_row(writer: Any, values: Iterable[Any]) -> None:
    writer.writerow([csv_safe_cell(value) for value in values])


def session_report(
    connection: sqlite3.Connection,
    home: Path,
    session: sqlite3.Row,
    items: Sequence[sqlite3.Row],
    score: float,
    answered: int,
) -> tuple[Path, Path]:
    stamp = utc_now().strftime("%Y%m%d-%H%M%S")
    report_dir = home / "reports" / utc_now().strftime("%Y") / utc_now().strftime("%m")
    report_dir.mkdir(parents=True, exist_ok=True)
    md_path = report_dir / f"{stamp}-{session['kind']}-{session['id'][:8]}.md"
    csv_path = report_dir / f"{stamp}-{session['kind']}-{session['id'][:8]}.csv"

    labels = {"daily": "每日复习", "weekly": "周测", "monthly": "月测"}
    scheduled = len(items)
    wrong_items = [item for item in items if item["first_correct"] == 0]
    exhausted = [item for item in items if item["status"] == "remediation_exhausted"]
    remediated = [item for item in wrong_items if item["status"] == "done"]
    weak_counts: dict[tuple[str, str], int] = defaultdict(int)
    for item in wrong_items:
        weak_counts[(item["subject"], item["topic"])] += 1

    lines = [
        f"# {labels.get(session['kind'], session['kind'])}报告",
        "",
        f"- 开始时间：{local_display(session['started_at_utc'])}",
        f"- 结束时间：{local_display(iso_utc())}",
        f"- 首次作答：{answered}/{scheduled}",
        f"- 首次正确率：{score:.1f}%",
        f"- 首次错题：{len(wrong_items)}",
        f"- 已完成补救：{len(remediated)}/{len(wrong_items)}",
        f"- 达到补救上限：{len(exhausted)}",
        "",
        "## 明细",
        "",
        "_以下题目和答案均作为已转义的数据显示，不执行 Markdown 或 HTML。_",
        "",
        "| 学科 | 知识点 | 题目 | 标准答案 | 首次结果 | 补救次数 | 状态 |",
        "|---|---|---|---|---:|---:|---|",
    ]
    for item in items:
        detail_template = (
            "| {subject} | {topic} | {prompt} | {answer} | {result} | {attempts} | {status} |"
        )
        lines.append(
            detail_template.format(
                subject=markdown_escape(item["subject"]),
                topic=markdown_escape(item["topic"]),
                prompt=markdown_escape(item["prompt"]),
                answer=markdown_escape(item["answer"]),
                result=markdown_escape(item["first_result"] or "未作答"),
                attempts=item["remediation_attempts"],
                status=markdown_escape(item["status"]),
            )
        )
    lines.extend(["", "## 薄弱知识点", ""])
    if weak_counts:
        for (subject, topic), count in sorted(
            weak_counts.items(), key=lambda pair: (-pair[1], pair[0])
        ):
            lines.append(
                f"- {markdown_escape(subject)} / {markdown_escape(topic)}：首次答错 {count} 题"
            )
    else:
        lines.append("- 本次没有首次答错的知识点。")
    lines.extend(
        [
            "",
            "## 下一步",
            "",
            "- 按原 FSRS 到期计划继续复习；补救答对不会覆盖首次错误记录。",
            "- 下次优先处理逾期和预计记忆率较低的内容。",
            "",
        ]
    )
    atomic_write_text(md_path, "\n".join(lines))

    stream = io.StringIO(newline="")
    writer = csv.writer(stream, lineterminator="\n")
    write_csv_row(
        writer,
        [
            "subject",
            "topic",
            "type",
            "prompt",
            "answer",
            "first_result",
            "first_correct",
            "remediation_attempts",
            "status",
        ],
    )
    for item in items:
        write_csv_row(
            writer,
            [
                item["subject"],
                item["topic"],
                item["item_type"],
                item["prompt"],
                item["answer"],
                item["first_result"] or "",
                "" if item["first_correct"] is None else item["first_correct"],
                item["remediation_attempts"],
                item["status"],
            ],
        )
    atomic_write_text(csv_path, "\ufeff" + stream.getvalue())
    return md_path, csv_path


def finish_session(home: Path, config: dict[str, Any], session_id: str) -> dict[str, Any]:
    connection = connect_db(home)
    try:
        begin_immediate(connection)
        session = connection.execute(
            "SELECT * FROM sessions WHERE id = ?", (session_id,)
        ).fetchone()
        if session is None:
            raise StudyError("学习会话不存在。")
        if session["status"] != "active":
            return {
                "session_id": session_id,
                "status": session["status"],
                "score_first": session["score_first"],
                "report_md": session["report_md"],
                "report_csv": session["report_csv"],
                "already_finished": True,
            }
        unresolved = connection.execute(
            """
            SELECT COUNT(*) FROM session_items
            WHERE session_id = ? AND status = 'remediation'
            """,
            (session_id,),
        ).fetchone()[0]
        if unresolved:
            raise StudyError(f"还有 {unresolved} 道错题未补救到答对，暂不能结束。")
        items = connection.execute(
            """
            SELECT si.*, i.subject, i.topic, i.item_type, i.prompt, i.answer
            FROM session_items si JOIN items i ON i.id = si.item_id
            WHERE si.session_id = ? ORDER BY si.ordinal
            """,
            (session_id,),
        ).fetchall()
        answered = sum(1 for item in items if item["first_result"] is not None)
        correct = sum(1 for item in items if item["first_correct"] == 1)
        score = (correct / answered * 100.0) if answered else 0.0
        md_path, csv_path = session_report(connection, home, session, items, score, answered)
        connection.execute(
            """
            UPDATE sessions
            SET ended_at_utc = ?, status = 'completed', score_first = ?,
                total_answered = ?, report_md = ?, report_csv = ?
            WHERE id = ?
            """,
            (
                iso_utc(),
                score,
                answered,
                str(md_path),
                str(csv_path),
                session_id,
            ),
        )
        connection.commit()
        dashboard = generate_dashboard(connection, home, config, write=True)
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    backup = backup_database(home, f"session-{session['kind']}", config)
    return {
        "session_id": session_id,
        "status": "completed",
        "kind": session["kind"],
        "scheduled": len(items),
        "answered": answered,
        "score_first": round(score, 2),
        "report_md": str(md_path),
        "report_csv": str(csv_path),
        "dashboard": dashboard["report"],
        "backup": str(backup),
    }


def test_status_data(connection: sqlite3.Connection, now: datetime) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for kind, interval in (("weekly", 7), ("monthly", 30)):
        last = connection.execute(
            """
            SELECT ended_at_utc, score_first, report_md
            FROM sessions
            WHERE kind = ? AND status = 'completed'
            ORDER BY ended_at_utc DESC LIMIT 1
            """,
            (kind,),
        ).fetchone()
        if last is None:
            output[kind] = {
                "due": True,
                "last_local": None,
                "last_score": None,
                "days_since": None,
            }
            continue
        last_dt = parse_dt(last["ended_at_utc"]) or now
        current_date = now.astimezone(LOCAL_TZ).date()
        last_date = last_dt.astimezone(LOCAL_TZ).date()
        elapsed = max(0, (current_date - last_date).days)
        output[kind] = {
            "due": elapsed >= interval,
            "last_local": local_display(last_dt),
            "last_score": last["score_first"],
            "days_since": elapsed,
            "report": last["report_md"],
        }
    return output


def generate_dashboard(
    connection: sqlite3.Connection,
    home: Path,
    config: dict[str, Any],
    write: bool,
) -> dict[str, Any]:
    now = utc_now()
    total = connection.execute("SELECT COUNT(*) FROM items WHERE active = 1").fetchone()[0]
    due_now = connection.execute(
        "SELECT COUNT(*) FROM items WHERE active = 1 AND due_utc <= ?",
        (iso_utc(now),),
    ).fetchone()[0]
    due_today = connection.execute(
        "SELECT COUNT(*) FROM items WHERE active = 1 AND due_utc <= ?",
        (iso_utc(end_of_local_day_utc(now)),),
    ).fetchone()[0]
    pending = connection.execute(
        "SELECT COUNT(*) FROM pending_imports WHERE status = 'pending'"
    ).fetchone()[0]
    unresolved = connection.execute("SELECT COUNT(*) FROM mistakes WHERE resolved = 0").fetchone()[
        0
    ]
    review_stats = connection.execute(
        """
        SELECT COUNT(*) AS total,
               COALESCE(SUM(CASE WHEN first_attempt_correct = 1 THEN 1 ELSE 0 END), 0)
                   AS correct
        FROM reviews WHERE is_remediation = 0
        """
    ).fetchone()
    review_total = int(review_stats["total"])
    accuracy = float(review_stats["correct"]) / review_total * 100.0 if review_total else None
    subjects = connection.execute(
        """
        SELECT subject, COUNT(*) AS total,
               SUM(CASE WHEN due_utc <= ? THEN 1 ELSE 0 END) AS due_today
        FROM items WHERE active = 1
        GROUP BY subject ORDER BY subject
        """,
        (iso_utc(end_of_local_day_utc(now)),),
    ).fetchall()
    weak = connection.execute(
        """
        SELECT i.subject, i.topic, COUNT(*) AS wrong_count
        FROM mistakes m JOIN items i ON i.id = m.item_id
        GROUP BY i.subject, i.topic
        ORDER BY wrong_count DESC, i.subject, i.topic LIMIT 10
        """
    ).fetchall()
    tests = test_status_data(connection, now)
    report_path = home / "reports" / "学习面板.md"
    lines = [
        "# 学习面板",
        "",
        f"更新时间：{local_display(now)}",
        "",
        "## 今日",
        "",
        f"- 已逾期/现在到期：{due_now}",
        f"- 今天结束前到期：{due_today}",
        f"- 待确认导入：{pending}",
        f"- 尚未补救完成的错题：{unresolved}",
        "",
        "## 总体",
        "",
        f"- 有效学习条目：{total}",
        f"- 首次作答次数：{review_total}",
        f"- 首次正确率：{'暂无' if accuracy is None else f'{accuracy:.1f}%'}",
        "",
        "## 分学科",
        "",
        "| 学科 | 条目 | 今天到期 |",
        "|---|---:|---:|",
    ]
    for row in subjects:
        lines.append(f"| {markdown_escape(row['subject'])} | {row['total']} | {row['due_today']} |")
    if not subjects:
        lines.append("| 暂无 | 0 | 0 |")
    lines.extend(["", "## 测试状态", ""])
    for kind, label in (("weekly", "周测"), ("monthly", "月测")):
        state = tests[kind]
        if state["last_local"] is None:
            lines.append(f"- {label}：尚未进行，当前建议开始。")
        else:
            due_text = "已到期" if state["due"] else "未到期"
            lines.append(
                f"- {label}：{due_text}；上次 {state['last_local']}，"
                f"首次得分 {state['last_score']:.1f}%"
            )
    lines.extend(["", "## 薄弱知识点", ""])
    if weak:
        for row in weak:
            lines.append(
                f"- {markdown_escape(row['subject'])} / {markdown_escape(row['topic'])}"
                f"：累计首次答错 {row['wrong_count']} 次"
            )
    else:
        lines.append("- 暂无错题记录。")
    lines.append("")
    if write:
        atomic_write_text(report_path, "\n".join(lines))
    return {
        "updated_local": local_display(now),
        "total_items": total,
        "due_now": due_now,
        "due_today": due_today,
        "pending": pending,
        "unresolved_mistakes": unresolved,
        "first_attempts": review_total,
        "first_accuracy": None if accuracy is None else round(accuracy, 2),
        "subjects": [dict(row) for row in subjects],
        "weak_topics": [dict(row) for row in weak],
        "tests": tests,
        "report": str(report_path),
    }


def dashboard(home: Path, config: dict[str, Any]) -> dict[str, Any]:
    connection = connect_db(home)
    try:
        return generate_dashboard(connection, home, config, write=True)
    finally:
        connection.close()


def test_status(home: Path) -> dict[str, Any]:
    connection = connect_db(home)
    try:
        return test_status_data(connection, utc_now())
    finally:
        connection.close()


def health(home: Path, workspace: Path) -> dict[str, Any]:
    initialized = db_path(home).exists()
    result: dict[str, Any] = {
        "healthy": True,
        "initialized": initialized,
        "confined_to_workspace": home.is_relative_to(workspace) and home != workspace,
        "workspace": str(workspace),
        "home": str(home),
        "database": str(db_path(home)),
        "external_pdf_process_disabled": bool(
            DEPENDENCY_IMPORT_ERROR is None
            and "pypdf" in globals()
            and pypdf.filters.JBIG2DEC_BINARY is None
        ),
        "scheduler": {
            "implementation": "Py-FSRS",
            "self_implemented": False,
            "desired_retention": DEFAULT_CONFIG["desired_retention"],
        },
        "packages": {},
    }
    for package in (
        "fsrs",
        "openpyxl",
        "pypdf",
        "et-xmlfile",
        "typing-extensions",
        "rapidocr",
        "onnxruntime",
        "opencv-python",
        "pillow",
        "numpy",
    ):
        try:
            result["packages"][package] = metadata.version(package)
        except metadata.PackageNotFoundError:
            result["packages"][package] = None
            result["healthy"] = False
    if DEPENDENCY_IMPORT_ERROR is not None:
        result["healthy"] = False
        result["startup_error"] = {
            "error_type": type(DEPENDENCY_IMPORT_ERROR).__name__,
            "error": str(DEPENDENCY_IMPORT_ERROR),
        }
    if not initialized:
        result["action"] = "run init"
        return result

    connection = connect_db(home)
    try:
        integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
        schema = connection.execute(
            "SELECT value FROM meta WHERE key = 'schema_version'"
        ).fetchone()
        result["integrity"] = integrity
        result["schema_version"] = schema[0] if schema else None
        result["counts"] = {
            "items": connection.execute("SELECT COUNT(*) FROM items").fetchone()[0],
            "reviews": connection.execute("SELECT COUNT(*) FROM reviews").fetchone()[0],
            "pending": connection.execute(
                "SELECT COUNT(*) FROM pending_imports WHERE status='pending'"
            ).fetchone()[0],
            "active_sessions": connection.execute(
                "SELECT COUNT(*) FROM sessions WHERE status='active'"
            ).fetchone()[0],
        }
        if integrity != "ok" or result["schema_version"] != SCHEMA_VERSION:
            result["healthy"] = False
    finally:
        connection.close()
    backups = sorted((home / "backups").glob("*.sqlite3"))
    result["backup_count"] = len(backups)
    result["latest_backup"] = str(backups[-1]) if backups else None
    result["backup_retention"] = {
        "count_limit": int(load_config(home).get("backup_retention_count", 30)),
        "max_total_mb": float(load_config(home).get("backup_max_total_mb", 2048)),
    }
    if backups:
        try:
            result["latest_backup_verification"] = verify_backup(backups[-1])
            if (
                result["latest_backup_verification"]["sqlite_integrity"] != "ok"
                or not result["latest_backup_verification"]["sha256_matches"]
            ):
                result["healthy"] = False
        except (OSError, sqlite3.Error, ValueError) as exc:
            result["latest_backup_verification"] = {
                "error_type": type(exc).__name__,
                "error": str(exc),
            }
            result["healthy"] = False
    runtime = home / "runtime" / ".venv"
    result["runtime_under_home"] = runtime.resolve().is_relative_to(home)
    return result


def emit(data: Any) -> None:
    print(json.dumps(data, ensure_ascii=False, indent=2, default=str))


def add_common_import_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--file", required=True, help="导入文件路径")
    parser.add_argument("--default-subject", default="英语")
    parser.add_argument("--default-topic", default="未分类")
    parser.add_argument("--confidence", type=float, default=0.95)


def build_parser() -> argparse.ArgumentParser:
    parser = JsonArgumentParser(description="自适应学习复习数据管理器")
    parser.add_argument("--home", help="学习系统目录；必须位于当前仓库工作区")
    subparsers = parser.add_subparsers(dest="command", required=True)

    subparsers.add_parser("init", help="初始化数据库与目录")
    subparsers.add_parser("health", help="检查路径、依赖与数据库")
    subparsers.add_parser("dashboard", help="更新并读取学习面板")
    subparsers.add_parser("test-status", help="查看周测/月测是否到期")
    subparsers.add_parser("pending", help="列出待确认导入")

    import_parser = subparsers.add_parser("import", help="导入学习内容")
    add_common_import_arguments(import_parser)

    image_parser = subparsers.add_parser("image-prepare", help="离线增强并识别照片、截图或扫描图像")
    image_parser.add_argument("--file", required=True, help="仓库工作区中的图像路径")
    image_parser.add_argument("--force", action="store_true", help="忽略缓存并重新处理")

    resolve_parser = subparsers.add_parser("pending-resolve", help="确认待处理条目")
    resolve_parser.add_argument("--pending-id", required=True)
    resolve_parser.add_argument("--subject")
    resolve_parser.add_argument("--topic")
    resolve_parser.add_argument("--type")
    resolve_parser.add_argument("--prompt")
    resolve_parser.add_argument("--answer", required=True)

    start_parser = subparsers.add_parser("session-start", help="开始复习或测试")
    start_parser.add_argument("--kind", choices=["daily", "weekly", "monthly"], required=True)
    start_parser.add_argument("--minutes", type=int)
    start_parser.add_argument("--limit", type=int)
    start_parser.add_argument("--new", action="store_true", help="忽略今天已有的活动会话")

    show_parser = subparsers.add_parser("session-show", help="读取会话与题目")
    show_parser.add_argument("--session-id", required=True)

    answer_parser = subparsers.add_parser("answer", help="记录首次答案或补救变式")
    answer_parser.add_argument("--session-id", required=True)
    answer_parser.add_argument("--item-id", required=True)
    answer_parser.add_argument(
        "--result", choices=["again", "hard", "good", "easy", "wrong", "correct"], required=True
    )
    answer_parser.add_argument("--answer-text", default="")
    answer_parser.add_argument("--reason")
    answer_parser.add_argument("--response-ms", type=int)
    answer_parser.add_argument("--remediation", action="store_true")

    finish_parser = subparsers.add_parser("session-finish", help="结束会话并生成报告")
    finish_parser.add_argument("--session-id", required=True)

    backup_parser = subparsers.add_parser("backup", help="创建完整 SQLite 备份")
    backup_parser.add_argument("--reason", default="manual")
    return parser


def run(args: argparse.Namespace) -> dict[str, Any]:
    workspace, home = resolve_home(args.home)
    if args.command == "health":
        return health(home, workspace)
    if DEPENDENCY_IMPORT_ERROR is not None:
        raise StudyError(
            "运行依赖不可用；请执行仓库内 setup 脚本后重试。"
            f"（{type(DEPENDENCY_IMPORT_ERROR).__name__}）"
        )
    if args.command == "init":
        return initialize(home)

    ensure_dirs(home)
    config = load_config(home)
    if args.command == "import":
        return import_file(
            home,
            config,
            resolve_workspace_input(args.file),
            args.default_subject,
            args.default_topic,
            args.confidence,
        )
    if args.command == "image-prepare":
        from prepare_scan import prepare_scan

        return prepare_scan(str(home), args.file, args.force)
    if args.command == "pending":
        return pending_list(home)
    if args.command == "pending-resolve":
        return pending_resolve(home, config, args)
    if args.command == "session-start":
        return start_session(home, config, args.kind, args.minutes, args.limit, args.new)
    if args.command == "session-show":
        return show_session(home, config, args.session_id)
    if args.command == "answer":
        return record_answer(
            home,
            config,
            args.session_id,
            args.item_id,
            args.result,
            args.answer_text,
            args.reason,
            args.response_ms,
            args.remediation,
        )
    if args.command == "session-finish":
        return finish_session(home, config, args.session_id)
    if args.command == "dashboard":
        return dashboard(home, config)
    if args.command == "test-status":
        return test_status(home)
    if args.command == "backup":
        return {"backup": str(backup_database(home, args.reason, config))}
    raise StudyError(f"未知命令：{args.command}")


def main() -> int:
    try:
        parser = build_parser()
        args = parser.parse_args()
        emit(run(args))
        return 0
    except StudyError as exc:
        emit({"ok": False, "error": str(exc)})
        return 2
    except (OSError, sqlite3.Error, json.JSONDecodeError, ValueError) as exc:
        emit({"ok": False, "error": str(exc), "error_type": type(exc).__name__})
        return 3
    except Exception as exc:  # Fail closed and preserve a machine-readable response.
        emit(
            {
                "ok": False,
                "error": "管理器发生未预期错误；数据事务已回滚，请运行 health。",
                "error_type": type(exc).__name__,
                "action": "run health",
            }
        )
        return 4


if __name__ == "__main__":
    sys.exit(main())
